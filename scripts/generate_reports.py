"""
Marketplace Sales Reporting — Automated Spreadsheet Generator
Generates all 5 required outputs from weekly/daily CSV exports
Run this script every Monday AM (or schedule via Task Scheduler on Windows)
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

# ── CONFIG — Update these paths before running ──────────────────────────────
WEEKLY_DATA_PATH = r"C:\PowerBI_Marketplace\Data\Weekly\1_Weekly_Sales.xlsx"
DAILY_DATA_PATH  = r"C:\PowerBI_Marketplace\Data\Daily\2_Daily_Sales.xlsx"
OUTPUT_FOLDER    = r"C:\PowerBI_Marketplace\Outputs"
TODAY            = datetime.today()
CURRENT_WEEK     = TODAY.isocalendar()[1]
# ────────────────────────────────────────────────────────────────────────────

# ── STYLING CONSTANTS ────────────────────────────────────────────────────────
NAVY   = "1F3864"
GOLD   = "C9A961"
WHITE  = "FFFFFF"
LGREY  = "F2F0EC"
GREEN  = "E2EFDA"
FONT   = "Arial"

def hdr_fill(): return PatternFill("solid", start_color=NAVY)
def gold_fill(): return PatternFill("solid", start_color=GOLD)
def grey_fill(): return PatternFill("solid", start_color=LGREY)
def grn_fill():  return PatternFill("solid", start_color=GREEN)
def hdr_font(sz=10): return Font(bold=True, color=WHITE, name=FONT, size=sz)
def body_font(sz=9): return Font(name=FONT, size=sz)
def bold_font(sz=9): return Font(name=FONT, size=sz, bold=True)
def border():
    s = Side(style="thin", color="D0D7E3")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row=1, height=26):
    for cell in ws[row]:
        cell.font = hdr_font()
        cell.fill = hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border()
    ws.row_dimensions[row].height = height

def style_data_rows(ws, start=2):
    for i, row in enumerate(ws.iter_rows(min_row=start, max_row=ws.max_row)):
        fill = grey_fill() if i % 2 == 0 else PatternFill("solid", start_color=WHITE)
        for cell in row:
            cell.font = body_font()
            cell.fill = fill
            cell.border = border()
            cell.alignment = Alignment(vertical="center")

def auto_width(ws, min_w=8, max_w=25):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)

def add_title_row(ws, title, subtitle=""):
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color=NAVY, name=FONT, size=13)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    if subtitle:
        ws.merge_cells(f"A2:{get_column_letter(ws.max_column)}2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(color="888888", name=FONT, size=9, italic=True)
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 14

def save_wb(wb, filename):
    path = os.path.join(OUTPUT_FOLDER, filename)
    wb.save(path)
    print(f"  ✓ Saved: {filename}")
    return path

# ── LOAD DATA ────────────────────────────────────────────────────────────────
print("Loading data...")
weekly = pd.read_excel(WEEKLY_DATA_PATH)
daily  = pd.read_excel(DAILY_DATA_PATH)
weekly["week_commencing"] = pd.to_datetime(weekly["week_commencing"])
print(f"  Weekly rows: {len(weekly):,}   Daily rows: {len(daily):,}")

MAX_WEEK = weekly["week_number"].max()
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# ════════════════════════════════════════════════════════════════════════════
# OUTPUT 1 — Weekly Sales by Channel
# Simple weekly table: sales £ and units by channel per week
# Delivered: Monday AM
# ════════════════════════════════════════════════════════════════════════════
def output1_weekly_by_channel():
    print("\n[1] Weekly Sales by Channel...")
    grp = weekly.groupby(["week_number", "channel_name"]).agg(
        sales_value=("sales_value", "sum"),
        sales_units=("sales_units", "sum")
    ).reset_index()

    # Pivot: rows = channel, cols = weeks
    piv_val = grp.pivot(index="channel_name", columns="week_number", values="sales_value").fillna(0).round(2)
    piv_uni = grp.pivot(index="channel_name", columns="week_number", values="sales_units").fillna(0).astype(int)

    wb = Workbook()

    # Sheet 1: Sales £ by channel
    ws1 = wb.active
    ws1.title = "Sales £ by Channel"
    headers = ["Channel"] + [f"W{w} £" for w in piv_val.columns]
    ws1.append(headers)
    for ch in piv_val.index:
        row = [ch] + [piv_val.loc[ch, w] for w in piv_val.columns]
        ws1.append(row)
    # Totals row
    total_row = ["TOTAL"] + [piv_val[w].sum().round(2) for w in piv_val.columns]
    ws1.append(total_row)
    for cell in ws1[ws1.max_row]:
        cell.font = bold_font()
        cell.fill = gold_fill()
        cell.border = border()
    # Format £ columns
    for row in ws1.iter_rows(min_row=2, min_col=2, max_col=ws1.max_column):
        for cell in row:
            cell.number_format = "£#,##0.00"
    style_header_row(ws1)
    style_data_rows(ws1, start=2)
    auto_width(ws1)
    ws1.freeze_panes = "B2"
    add_title_row(ws1, "Weekly Sales by Channel — Sales £",
                  f"Generated: {TODAY.strftime('%d %b %Y')}  |  Weeks W1–W{MAX_WEEK}")

    # Sheet 2: Units by channel
    ws2 = wb.create_sheet("Units by Channel")
    headers2 = ["Channel"] + [f"W{w} Units" for w in piv_uni.columns]
    ws2.append(headers2)
    for ch in piv_uni.index:
        row = [ch] + [int(piv_uni.loc[ch, w]) for w in piv_uni.columns]
        ws2.append(row)
    total_row2 = ["TOTAL"] + [int(piv_uni[w].sum()) for w in piv_uni.columns]
    ws2.append(total_row2)
    for cell in ws2[ws2.max_row]:
        cell.font = bold_font()
        cell.fill = gold_fill()
        cell.border = border()
    style_header_row(ws2)
    style_data_rows(ws2, start=2)
    auto_width(ws2)
    ws2.freeze_panes = "B2"
    add_title_row(ws2, "Weekly Sales by Channel — Units",
                  f"Generated: {TODAY.strftime('%d %b %Y')}  |  Weeks W1–W{MAX_WEEK}")

    fname = f"Output1_Weekly_Sales_by_Channel_W{MAX_WEEK}.xlsx"
    save_wb(wb, fname)


# ════════════════════════════════════════════════════════════════════════════
# OUTPUT 2 — Sales by Style (Rolling 5 Weeks)
# Style-level view per brand, rolling 5-week window
# ════════════════════════════════════════════════════════════════════════════
def output2_sales_by_style():
    print("\n[2] Sales by Style — Rolling 5 Weeks...")
    last5 = sorted(weekly["week_number"].unique())[-5:]
    cw_labels = {w: f"CW" if i == len(last5)-1 else f"CW-{len(last5)-1-i}"
                 for i, w in enumerate(last5)}

    wb = Workbook()
    first = True

    for brand in sorted(weekly["brand"].unique()):
        df_brand = weekly[weekly["brand"] == brand].copy()
        ws = wb.active if first else wb.create_sheet(brand)
        ws.title = brand
        first = False

        # Build style summary
        style_grp = df_brand.groupby(["style_code", "description", "category", "season"]).agg(
            total_units=("sales_units", "sum"),
            total_value=("sales_value", "sum"),
            stock_units=("stock_units", "sum"),
            replen_units=("replenishment_units", "sum"),
            on_order=("on_order_units", "sum"),
            retail_price=("retail_price", "mean"),
        ).reset_index()

        # Rolling 5-week units per style
        for wk in last5:
            wk_data = df_brand[df_brand["week_number"] == wk].groupby("style_code")["sales_units"].sum()
            style_grp[cw_labels[wk]] = style_grp["style_code"].map(wk_data).fillna(0).astype(int)

        # ASP
        style_grp["ASP"] = (style_grp["total_value"] / style_grp["total_units"].replace(0, np.nan)).round(2)
        # Sell-Through %
        style_grp["Sell-Through %"] = (
            style_grp["total_units"] / (style_grp["total_units"] + style_grp["stock_units"])
        ).fillna(0).round(4)
        # Weeks Cover
        avg_wkly = style_grp["total_units"] / 5
        style_grp["Weeks Cover"] = (style_grp["stock_units"] / avg_wkly.replace(0, np.nan)).round(1)

        # Build output cols
        week_cols = [cw_labels[w] for w in last5]
        cols = ["style_code", "description", "category", "season"] + week_cols + \
               ["total_units", "ASP", "stock_units", "Weeks Cover", "Sell-Through %",
                "replen_units", "on_order"]
        headers = ["Style Code", "Description", "Category", "Season"] + \
                  [f"Units {c}" for c in week_cols] + \
                  ["Total Units", "ASP £", "Stock", "Weeks Cover", "Sell-Through %",
                   "Replen Units", "On Order"]

        ws.append(headers)
        for _, row in style_grp.sort_values("total_units", ascending=False).iterrows():
            ws.append([row[c] for c in cols])

        # Format
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
            for cell in row:
                cell.number_format = "£#,##0.00"
        for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
            for cell in row:
                cell.number_format = "0.0%"

        style_header_row(ws)
        style_data_rows(ws, start=2)
        auto_width(ws)
        ws.freeze_panes = "E2"
        add_title_row(ws, f"Sales by Style — {brand} — Rolling 5 Weeks",
                      f"CW = W{MAX_WEEK}  |  Generated: {TODAY.strftime('%d %b %Y')}")

    fname = f"Output2_Sales_by_Style_Rolling5wk_W{MAX_WEEK}.xlsx"
    save_wb(wb, fname)


# ════════════════════════════════════════════════════════════════════════════
# OUTPUT 3 — Channel-Specific by Style
# Same as Output 2 but filtered per channel, one sheet per channel
# ════════════════════════════════════════════════════════════════════════════
def output3_channel_by_style():
    print("\n[3] Channel-Specific Sales by Style...")
    last5 = sorted(weekly["week_number"].unique())[-5:]
    cw_labels = {w: "CW" if i == len(last5)-1 else f"CW-{len(last5)-1-i}"
                 for i, w in enumerate(last5)}

    wb = Workbook()
    first = True

    for channel in sorted(weekly["channel_name"].unique()):
        df_ch = weekly[weekly["channel_name"] == channel].copy()
        ws = wb.active if first else wb.create_sheet(channel[:28])
        ws.title = channel[:28]
        first = False

        style_grp = df_ch.groupby(["style_code", "description", "brand", "category"]).agg(
            total_units=("sales_units", "sum"),
            total_value=("sales_value", "sum"),
            stock_units=("stock_units", "sum"),
        ).reset_index()

        for wk in last5:
            wk_data = df_ch[df_ch["week_number"] == wk].groupby("style_code")["sales_units"].sum()
            style_grp[cw_labels[wk]] = style_grp["style_code"].map(wk_data).fillna(0).astype(int)

        style_grp["ASP"] = (style_grp["total_value"] / style_grp["total_units"].replace(0, np.nan)).round(2)

        week_cols = [cw_labels[w] for w in last5]
        cols = ["style_code", "description", "brand", "category"] + week_cols + ["total_units", "ASP"]
        headers = ["Style Code", "Description", "Brand", "Category"] + \
                  [f"Units {c}" for c in week_cols] + ["Total Units", "ASP £"]

        ws.append(headers)
        for _, row in style_grp.sort_values("total_units", ascending=False).iterrows():
            ws.append([row[c] for c in cols])

        for row in ws.iter_rows(min_row=2, min_col=ws.max_column, max_col=ws.max_column):
            for cell in row:
                cell.number_format = "£#,##0.00"

        style_header_row(ws)
        style_data_rows(ws, start=2)
        auto_width(ws)
        ws.freeze_panes = "E2"
        add_title_row(ws, f"Channel Report: {channel} — Sales by Style",
                      f"CW = W{MAX_WEEK}  |  Generated: {TODAY.strftime('%d %b %Y')}")

    fname = f"Output3_Channel_by_Style_W{MAX_WEEK}.xlsx"
    save_wb(wb, fname)


# ════════════════════════════════════════════════════════════════════════════
# OUTPUT 4 — Sales by EAN (Barcode Level)
# Granular barcode-level weekly data
# ════════════════════════════════════════════════════════════════════════════
def output4_sales_by_ean():
    print("\n[4] Sales by EAN (Barcode Level)...")
    df = weekly[[
        "week_number", "week_commencing", "channel_name", "style_code",
        "colour", "size", "EAN", "sales_units", "sales_value",
        "cost", "retail_price", "returns_units", "returns_value"
    ]].copy()

    df = df.sort_values(["week_number", "channel_name", "style_code", "size"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales by EAN"

    headers = [
        "Week No", "Week Commencing", "Channel", "Style Code",
        "Colour", "Size", "EAN", "Sales Units", "Sales Value £",
        "Cost £", "Retail Price £", "Returns Units", "Returns Value £"
    ]
    ws.append(headers)

    for _, row in df.iterrows():
        ws.append([
            row["week_number"],
            row["week_commencing"].strftime("%d/%m/%Y") if pd.notna(row["week_commencing"]) else "",
            row["channel_name"], row["style_code"],
            row["colour"], row["size"], str(row["EAN"]),
            int(row["sales_units"]), round(row["sales_value"], 2),
            round(row["cost"], 2), round(row["retail_price"], 2),
            int(row["returns_units"]), round(row["returns_value"], 2)
        ])

    # Format £ columns
    for col_idx in [9, 10, 11, 13]:
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = "£#,##0.00"

    style_header_row(ws)
    style_data_rows(ws, start=2)
    auto_width(ws)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    add_title_row(ws, f"Sales by EAN — Barcode Level Detail",
                  f"All weeks  |  Generated: {TODAY.strftime('%d %b %Y')}")

    fname = f"Output4_Sales_by_EAN_W{MAX_WEEK}.xlsx"
    save_wb(wb, fname)


# ════════════════════════════════════════════════════════════════════════════
# OUTPUT 5 — Daily Sales by Barcode
# One file per day, delivered pre-7am daily
# Fields: branch, date, barcode, sales_value, sales_units
# ════════════════════════════════════════════════════════════════════════════
def output5_daily_barcode():
    print("\n[5] Daily Sales by Barcode...")
    daily_copy = daily.copy()
    if "date" in daily_copy.columns:
        daily_copy["date"] = pd.to_datetime(daily_copy["date"])

    # Generate one file per date in the daily data
    dates = daily_copy["date"].unique() if "date" in daily_copy.columns else [TODAY]

    for sale_date in sorted(dates):
        if pd.isna(sale_date):
            continue
        date_str = pd.Timestamp(sale_date).strftime("%Y-%m-%d")
        df_day = daily_copy[daily_copy["date"] == sale_date].copy() if "date" in daily_copy.columns else daily_copy

        wb = Workbook()
        ws = wb.active
        ws.title = f"Daily Sales {date_str}"

        headers = ["Branch", "Date", "Barcode", "Sales Value £", "Sales Units"]
        ws.append(headers)

        for _, row in df_day.iterrows():
            ws.append([
                str(row.get("branch", row.get("StoreID", ""))),
                pd.Timestamp(sale_date).strftime("%d/%m/%Y"),
                str(row.get("barcode", row.get("EAN", ""))),
                round(float(row.get("sales_value", row.get("SalesValue", 0))), 2),
                int(row.get("sales_units", row.get("SalesUnits", 0)))
            ])

        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = "£#,##0.00"

        style_header_row(ws)
        style_data_rows(ws, start=2)
        auto_width(ws)
        ws.freeze_panes = "B2"
        add_title_row(ws, f"Daily Sales by Barcode — {pd.Timestamp(sale_date).strftime('%d %B %Y')}",
                      f"Generated: {TODAY.strftime('%d %b %Y %H:%M')}  |  Pre-7am delivery")

        fname = f"Output5_Daily_Barcode_{date_str}.xlsx"
        save_wb(wb, fname)


# ── RUN ALL OUTPUTS ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  MARKETPLACE SALES REPORTING — AUTO GENERATOR")
    print(f"  Run date: {TODAY.strftime('%A %d %B %Y %H:%M')}")
    print(f"  Current week: W{CURRENT_WEEK}")
    print("=" * 60)

    output1_weekly_by_channel()
    output2_sales_by_style()
    output3_channel_by_style()
    output4_sales_by_ean()
    output5_daily_barcode()

    print("\n" + "=" * 60)
    print("  ALL REPORTS GENERATED SUCCESSFULLY!")
    print(f"  Location: {OUTPUT_FOLDER}")
    print("=" * 60)
